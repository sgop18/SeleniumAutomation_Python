import openpyxl
from openpyxl import workbook


        
def getRowCount(excelpath,sheetName):
    workbook=openpyxl.load_workbook(excelpath)
    sheet = workbook[sheetName]
    row=sheet.max_row
    return row

def getColCount(excelpath,sheetName):
    workbook=openpyxl.load_workbook(excelpath)
    sheet = workbook[sheetName]
    col=sheet.max_column
    return col

def readData(excelpath,sheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(excelpath)
    sheet = workbook[sheetName]
    row=rownum
    col=colnum
    cellValue=sheet.cell(row,col).value
    return cellValue

def writeData(excelpath,sheetName,rownum,colnum,data):
    workbook=openpyxl.load_workbook(excelpath)
    sheet = workbook[sheetName]
    row=rownum
    col=colnum
    sheet.cell(row,col).value = data
    workbook.save(excelpath)
    
    
